import io
import openpyxl
from src.utils.excel_kyc_generator import generar_excel_kyc_corporativo
from src.utils.excel_kyc_parser import parse_excel_kyc_file

excel_bytes = generar_excel_kyc_corporativo("Nelson Orlando Moraga Benavides", True, True, True)

wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
ws = wb.active

ws['D5'] = "Matilde Ivette Rivera Berndt (Cónyuge; RUT 9.017.656-0), Nelson Rodrigo Moraga Cires (Hijo, RUT 10.870.820-4)"
ws['D6'] = "ROL 01026-0001 Peñuelas-Coquimbo, Valor comercial aprox. $150.000.000. Sin deuda hipotecaria."
ws['D7'] = "No sé cuando tengo en PRINCIPAL o en APV Régimen B (a lo mejor tengo algo en el Banco Chile por mis cuentas corrientes)."

out_buf = io.BytesIO()
wb.save(out_buf)
filled_bytes = out_buf.getvalue()

res = parse_excel_kyc_file(filled_bytes)
print("--- RESULTADOS EXTRAÍDOS DEL EXCEL ---")
print("\nHerederos:")
for h in res["herederos"]:
    print(" ", h)

print("\nPropiedades:")
for p in res["propiedades"]:
    print(" ", p)

print("\nPólizas / APV:")
for pol in res["polizas"]:
    print(" ", pol)
