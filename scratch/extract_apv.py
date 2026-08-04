import pandas as pd
import openpyxl

path = r"C:\Users\franc\OneDrive\Documentos\PROYECTOS\BD SENIOR\SIMULADORES Y OTROS\Simulador Calculadora Ahorro APV y Pension.xlsm"

wb = openpyxl.load_workbook(path, data_only=False)
print("Hojas:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Hoja: {sheet_name} ---")
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=10):
        for cell in row:
            if cell.value is not None:
                # solo imprimir celdas que parecen contener texto descriptivo o formulas
                if isinstance(cell.value, str) and (cell.value.startswith('=') or len(cell.value) > 3):
                    print(f"{cell.coordinate}: {cell.value}")
