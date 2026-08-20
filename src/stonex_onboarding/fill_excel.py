import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TEMPLATE_PATH = "STONEX/PN - Datos apertura de cuenta Stonex (1).xlsx"

def safe_write(sheet, coord, value):
    """Escribe un valor en la celda, manejando el caso de que sea un MergedCell."""
    cell = sheet[coord]
    if type(cell).__name__ == 'MergedCell':
        for merged_range in sheet.merged_cells.ranges:
            if coord in merged_range:
                top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left.value = value
                return
    else:
        cell.value = value

def generate_stonex_excel(data, output_filename="Onboarding_Stonex.xlsx"):
    """
    Rellena la plantilla de Stonex con los datos del formulario web.
    """
    tipo_cuenta = data.get('tipo_cuenta', 'Persona Natural')
    
    if tipo_cuenta == 'Persona Jurídica':
        template_path = "STONEX/PJ - Datos apertura de cuenta Stonex.xlsx"
    else:
        template_path = "STONEX/PN - Datos apertura de cuenta Stonex (1).xlsx"
        
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"No se encontró la plantilla en {template_path}")

    wb = load_workbook(template_path)
    sheet = wb['Datos - solicitar a cliente']

    if tipo_cuenta == 'Persona Jurídica':
        # Perfilamiento de Riesgo (PJ)
        safe_write(sheet, 'D31', data.get('horizonte_tiempo', ''))
        safe_write(sheet, 'D33', data.get('experiencia', ''))
        safe_write(sheet, 'D35', data.get('porcentaje_activos', ''))
        safe_write(sheet, 'D37', data.get('objetivos_inversion', ''))
        safe_write(sheet, 'D39', data.get('tolerancia_riesgo', ''))
        safe_write(sheet, 'D41', data.get('eleccion_portafolio', ''))
        safe_write(sheet, 'D43', data.get('eleccion_rendimiento', ''))
        safe_write(sheet, 'D46', data.get('caida_portafolio', ''))

        # Mapeo de Empresa
        safe_write(sheet, 'B55', data.get('rut_empresa', ''))
        safe_write(sheet, 'B56', data.get('nombre_completo', ''))
        safe_write(sheet, 'B57', data.get('rubro', ''))
        safe_write(sheet, 'B58', data.get('actividad_economica', ''))
        safe_write(sheet, 'B59', data.get('direccion_empresa', ''))
        safe_write(sheet, 'B60', data.get('pais_empresa', 'Chile'))
        safe_write(sheet, 'B61', data.get('provincia_empresa', ''))
        safe_write(sheet, 'B62', data.get('ciudad_empresa', ''))
        safe_write(sheet, 'B63', data.get('cp_emp', ''))
        safe_write(sheet, 'B64', data.get('telefono_empresa', ''))
        safe_write(sheet, 'B65', data.get('email_empresa', ''))
        safe_write(sheet, 'B66', data.get('descripcion_inversiones', ''))
        safe_write(sheet, 'B67', data.get('acepta_correo', 'Sí'))
        safe_write(sheet, 'B68', data.get('deposito_terceros', 'No'))
        safe_write(sheet, 'B69', data.get('fondo_cobertura', 'No'))
        safe_write(sheet, 'B70', data.get('negocios_us', 'No'))
        safe_write(sheet, 'B71', data.get('cuentas_us', 'No'))
        safe_write(sheet, 'B72', data.get('acciones_portador', 'No'))
        safe_write(sheet, 'B73', data.get('es_banco', 'No'))
        safe_write(sheet, 'B74', data.get('es_broker', 'No'))
        safe_write(sheet, 'B75', data.get('casa_cambio', 'No'))
        safe_write(sheet, 'B76', data.get('gubernamental', 'No'))
        safe_write(sheet, 'B77', data.get('fondo_asesor', 'No'))
        
        # Mapeo de Experiencia PJ
        safe_write(sheet, 'B80', data.get('exp_acciones', ''))
        safe_write(sheet, 'B81', data.get('exp_fondos', ''))
        safe_write(sheet, 'B82', data.get('exp_anualidades', ''))
        safe_write(sheet, 'B83', data.get('exp_opciones', ''))
        safe_write(sheet, 'B84', data.get('exp_alternativas', ''))
        
        # Mapeo Financiero y Perfilamiento PJ
        safe_write(sheet, 'B87', data.get('necesidad_liquidez', ''))
        safe_write(sheet, 'B88', data.get('ingresos_anuales', ''))
        safe_write(sheet, 'B89', data.get('ingresos_exacto', ''))
        safe_write(sheet, 'B90', data.get('activos_liquidos', ''))
        safe_write(sheet, 'B91', data.get('activos_exacto', ''))
        safe_write(sheet, 'B92', data.get('patrimonio_total', ''))
        safe_write(sheet, 'B93', data.get('aportes_adicionales', 'No'))
        safe_write(sheet, 'B94', data.get('tiempo_retiros', ''))
        safe_write(sheet, 'B95', data.get('banco_origen', ''))
        safe_write(sheet, 'B96', data.get('pais_origen_fondos', 'Chile'))
        safe_write(sheet, 'B97', data.get('banco_origen', ''))
        safe_write(sheet, 'B98', data.get('origen_fondos', ''))
        
        # Mapeo de Beneficiario/Representante Titular 1
        safe_write(sheet, 'B107', data.get('porcentaje_part', '100'))
        safe_write(sheet, 'B108', data.get('categoria_rep', 'Beneficiario Final'))
        safe_write(sheet, 'B109', data.get('nombre_rep', ''))
        safe_write(sheet, 'B110', data.get('apellido_rep', ''))
        safe_write(sheet, 'B111', data.get('dir_rep', ''))
        safe_write(sheet, 'B112', data.get('prov_rep', ''))
        safe_write(sheet, 'B113', data.get('ciud_rep', ''))
        safe_write(sheet, 'B114', data.get('cp_rep', ''))
        safe_write(sheet, 'B115', data.get('telefono_rep', ''))
        safe_write(sheet, 'B116', data.get('email_rep', ''))
        safe_write(sheet, 'B117', data.get('fecha_nac_rep', ''))
        safe_write(sheet, 'B118', data.get('pais_emisor_doc', 'Chile'))
        safe_write(sheet, 'B119', data.get('tipo_doc', ''))
        safe_write(sheet, 'B120', data.get('rut_rep', ''))
        safe_write(sheet, 'B121', data.get('fecha_emi_doc', ''))
        safe_write(sheet, 'B122', data.get('fecha_exp_doc', ''))
        safe_write(sheet, 'B124', data.get('nacionalidad', 'Chile'))
        safe_write(sheet, 'B125', data.get('ciudadania', 'Chile'))
        safe_write(sheet, 'B126', data.get('situacion_laboral', ''))
        safe_write(sheet, 'B127', data.get('genero', ''))
        safe_write(sheet, 'B128', data.get('ocupacion', ''))
        safe_write(sheet, 'B129', data.get('estado_civil', ''))
        safe_write(sheet, 'B130', data.get('cantidad_hijos', ''))
        safe_write(sheet, 'B131', data.get('pep', 'No'))
        
        # Bancos Habituales
        safe_write(sheet, 'B137', data.get('banco_pais', 'Chile'))
        safe_write(sheet, 'B138', data.get('banco_ciudad', ''))
        safe_write(sheet, 'B139', data.get('banco_nombre', ''))
        safe_write(sheet, 'B140', data.get('banco_sucursal', ''))
        
        # Casillas Laborales del Representante Legal
        if data.get('situacion_laboral') in ['Empleado', 'Empleado/Independiente']:
            safe_write(sheet, 'D123', data.get('cargo_rep', ''))
            safe_write(sheet, 'D124', data.get('empresa_rep', ''))
            safe_write(sheet, 'D125', data.get('rubro_rep', ''))
            safe_write(sheet, 'D126', data.get('pais_emp_rep', 'Chile'))
            safe_write(sheet, 'D127', data.get('prov_emp_rep', ''))
            safe_write(sheet, 'D128', data.get('ciud_emp_rep', ''))
            safe_write(sheet, 'D129', data.get('dir_emp_rep', ''))
            safe_write(sheet, 'D131', data.get('tel_emp_rep', ''))
            safe_write(sheet, 'D132', data.get('email_emp_rep', ''))

        # Campos de inversión
        safe_write(sheet, 'B101', data.get('monto_inversion', ''))
        safe_write(sheet, 'B102', data.get('tipo_contrato', 'Canalización de órdenes'))
        safe_write(sheet, 'B103', data.get('fee_anual', '1.0%'))
        safe_write(sheet, 'B104', '0')
        safe_write(sheet, 'B105', data.get('metodo_pago', 'Transferencia'))
    else:
        # Mapeo de Persona Natural
        # Perfilamiento de Riesgo
        safe_write(sheet, 'D31', data.get('horizonte_tiempo', ''))
        safe_write(sheet, 'D33', data.get('experiencia', ''))
        safe_write(sheet, 'D35', data.get('porcentaje_activos', ''))
        safe_write(sheet, 'D37', data.get('objetivos_inversion', ''))
        safe_write(sheet, 'D39', data.get('tolerancia_riesgo', ''))
        safe_write(sheet, 'D41', data.get('eleccion_portafolio', ''))
        safe_write(sheet, 'D43', data.get('eleccion_rendimiento', ''))
        safe_write(sheet, 'D46', data.get('caida_portafolio', ''))
        
        # Datos de Cuenta y Asesor
        safe_write(sheet, 'B17', 'Individual')
        safe_write(sheet, 'B20', data.get('rep_code', 'Asesor FV'))
        
        # Datos Personales
        safe_write(sheet, 'B49', data.get('pn_primer_nombre', ''))
        safe_write(sheet, 'B50', data.get('pn_segundo_nombre', ''))
        safe_write(sheet, 'B51', data.get('pn_primer_apellido', ''))
        safe_write(sheet, 'B52', data.get('pn_segundo_apellido', ''))
        safe_write(sheet, 'B53', data.get('direccion_residencia', ''))
        safe_write(sheet, 'B54', data.get('pais_residencia', 'Chile'))
        safe_write(sheet, 'B55', data.get('provincia', ''))
        safe_write(sheet, 'B56', data.get('ciudad', ''))
        safe_write(sheet, 'B57', data.get('codigo_postal', ''))
        safe_write(sheet, 'B58', data.get('direccion_correspondencia', ''))
        safe_write(sheet, 'B59', data.get('telefono', ''))
        safe_write(sheet, 'B60', data.get('email', ''))
        safe_write(sheet, 'B61', data.get('fecha_nacimiento', ''))
        safe_write(sheet, 'B62', data.get('pais_emisor_doc', 'Chile'))
        safe_write(sheet, 'B63', data.get('tipo_documento', 'ID Nacional'))
        safe_write(sheet, 'B64', data.get('rut', ''))
        safe_write(sheet, 'B65', data.get('rut', ''))  # Identificación fiscal (rut)
        
        safe_write(sheet, 'B66', data.get('nacionalidad', 'Chilena'))
        safe_write(sheet, 'B67', data.get('ciudadania', 'Chilena'))
        safe_write(sheet, 'B68', data.get('situacion_laboral', ''))
        safe_write(sheet, 'B69', data.get('ocupacion', ''))
        safe_write(sheet, 'B70', data.get('estado_civil', ''))
        safe_write(sheet, 'B71', data.get('cantidad_hijos', '0'))
        safe_write(sheet, 'B72', data.get('pep', 'No'))
        safe_write(sheet, 'B74', data.get('acepta_correo', 'Sí'))
        
        if data.get('estado_civil') == 'Casado/a':
            safe_write(sheet, 'B77', data.get('con_nombres', ''))
            safe_write(sheet, 'B78', data.get('con_apellidos', ''))
            safe_write(sheet, 'E77', data.get('con_fecha_nac', ''))
            safe_write(sheet, 'B79', data.get('con_sit_lab', ''))
            safe_write(sheet, 'B80', data.get('con_nac', ''))
            safe_write(sheet, 'E78', data.get('con_pais_emi', ''))
            safe_write(sheet, 'E79', data.get('con_tipo_doc', ''))
            safe_write(sheet, 'E80', data.get('con_rut', ''))
            fill_green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            for cell_id in ['E77', 'E78', 'E79', 'E80']:
                sheet[cell_id].fill = fill_green
            
        # Experiencia de Inversiones
        safe_write(sheet, 'B84', data.get('exp_acciones', 'Nula'))
        safe_write(sheet, 'B85', data.get('exp_fondos', 'Nula'))
        safe_write(sheet, 'B86', data.get('exp_anualidades', 'Nula'))
        safe_write(sheet, 'B87', data.get('exp_opciones', 'Nula'))
        safe_write(sheet, 'B88', data.get('exp_alternativas', 'Nula'))
        
        # Datos financieros y de Inversión
        safe_write(sheet, 'B92', data.get('necesidad_liquidez', ''))
        safe_write(sheet, 'B93', data.get('ingresos_anuales', ''))
        safe_write(sheet, 'B94', data.get('ingresos_exacto', ''))
        safe_write(sheet, 'B95', data.get('activos_liquidos', ''))
        safe_write(sheet, 'B96', data.get('activos_exacto', ''))
        safe_write(sheet, 'B97', data.get('patrimonio_total', ''))
        safe_write(sheet, 'B98', data.get('aportes_adicionales', 'No'))
        safe_write(sheet, 'B101', data.get('tiempo_retiros', 'Más de 10 años'))
        
        safe_write(sheet, 'B102', data.get('banco_origen', ''))
        safe_write(sheet, 'B103', data.get('pais_origen_fondos', 'Chile'))
        safe_write(sheet, 'B104', data.get('origen_fondos', ''))
        
        # Campos Inversión Final y Contrato
        safe_write(sheet, 'A109', data.get('monto_inversion', ''))
        safe_write(sheet, 'B110', data.get('tipo_contrato', 'Canalización de órdenes'))
        safe_write(sheet, 'A111', data.get('fee_anual', '1.0%'))
        safe_write(sheet, 'A112', '0')
        safe_write(sheet, 'B113', data.get('metodo_pago', 'Transferencia'))
        
        # Bancos Habituales PN
        safe_write(sheet, 'B117', data.get('banco_pais', 'Chile'))
        safe_write(sheet, 'B118', data.get('banco_ciudad', ''))
        safe_write(sheet, 'B119', data.get('banco_nombre', ''))
        safe_write(sheet, 'B120', data.get('banco_sucursal', ''))
        
        # Campos de texto libre
        safe_write(sheet, 'B109', data.get('monto_inversion', ''))
        safe_write(sheet, 'B111', data.get('fee_anual', '1.0%'))
        safe_write(sheet, 'B112', '0')
        
    # Desproteger la hoja para que el asesor pueda editarla manualmente después de la descarga
    sheet.protection.sheet = False

    # Guardar en la carpeta data/onboarding
    os.makedirs("data/onboarding", exist_ok=True)
    output_path = os.path.join("data/onboarding", output_filename)
    wb.save(output_path)
    
    return output_path
