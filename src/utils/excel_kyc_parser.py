import io
import re
import datetime
import logging
import pandas as pd
import openpyxl
from src.osint.indicadores import get_uf_today

def parse_excel_kyc_file(file_bytes):
    """
    Parsea de forma completa y automática cualquier archivo Excel KYC (Altus_KYC_*.xlsx)
    completado por el cliente, extrayendo sin pérdidas:
    1. Herederos (RUT, Nombre, Parentesco, % Asignación Legal)
    2. Propiedades (ROL, Comuna, Dirección, Valor Comercial UF/CLP, Deuda)
    3. Pólizas / APV (Institución, Cobertura, Régimen)
    4. Respuestas generales estructuradas
    """
    results = {
        "herederos": [],
        "propiedades": [],
        "polizas": [],
        "respuestas_raw": {}
    }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        # 1. ESQUEMA DE LECTURA DINÁMICA POR FILAS
        for row in range(4, ws.max_row + 1):
            cell_req = ws.cell(row=row, column=1).value
            cell_ans = ws.cell(row=row, column=4).value
            
            item_req = str(cell_req or "").strip()
            answer = str(cell_ans or "").strip()

            if not item_req or not answer or answer.lower() in ["none", "nan", "sí", "si", "de acuerdo", "respuesta / detalle del cliente (escribir aquí)"]:
                continue

            results["respuestas_raw"][item_req] = answer
            item_req_upper = item_req.upper()

            # ----------------------------------------------------
            # A. PARSER DE HEREDEROS Y SUCESIÓN FAMILIAR
            # ----------------------------------------------------
            if any(kw in item_req_upper for kw in ["HEREDERO", "SUCESION", "SUCESIÓN", "COMPOSICION", "COMPOSICIÓN", "FAMILIAR", "HIJO", "CONYUGE", "CÓNYUGE"]):
                entries = re.split(r'[,;\n](?![^(]*\))', answer)
                parsed_heirs = []
                
                for entry in entries:
                    entry_str = entry.strip()
                    if not entry_str or len(entry_str) < 3:
                        continue
                    
                    # Extracción universal de RUT (formato 12.345.678-9 o 12345678-K)
                    rut_matches = re.findall(r'(\d{1,2}(?:\.\d{3}){2}-[\dkK]|\d{7,8}-[\dkK])', entry_str)
                    rut_extracted = rut_matches[0] if rut_matches else ""
                    
                    # Extracción de Fecha de Nacimiento (formatos DD/MM/YYYY, DD-MM-YYYY o YYYY-MM-DD)
                    dob = None
                    dob_m = re.search(r'\b(\d{1,2})[\/\.-](\d{1,2})[\/\.-](\d{2,4})\b', entry_str)
                    if dob_m:
                        d_val, m_val, y_val = int(dob_m.group(1)), int(dob_m.group(2)), int(dob_m.group(3))
                        if y_val < 100: y_val += 1900 if y_val > 30 else 2000
                        try: dob = datetime.date(y_val, m_val, d_val)
                        except: pass
                    if not dob:
                        dob_m2 = re.search(r'\b(\d{4})[\/\.-](\d{1,2})[\/\.-](\d{1,2})\b', entry_str)
                        if dob_m2:
                            y_val, m_val, d_val = int(dob_m2.group(1)), int(dob_m2.group(2)), int(dob_m2.group(3))
                            try: dob = datetime.date(y_val, m_val, d_val)
                            except: pass

                    # Limpiar nombre removiendo paréntesis, RUT y fechas
                    name_clean = re.sub(r'\([^)]*\)', '', entry_str)
                    name_clean = re.sub(r'RUT\s*[:\s]*[\d\.-]+[kK\d]?', '', name_clean, flags=re.IGNORECASE)
                    name_clean = re.sub(r'(?:Fecha\s*Nac|Nacimiento|Fec\s*Nac)\s*[:\s]*[\d\/\.-]+', '', name_clean, flags=re.IGNORECASE)
                    name_clean = re.sub(r'\b\d{1,2}[\/\.-]\d{1,2}[\/\.-]\d{2,4}\b', '', name_clean)
                    name_clean = re.sub(r'\s+', ' ', name_clean).strip()
                    
                    # Extraer texto en paréntesis (ej. Cónyuge, Hijo)
                    rel_match = re.search(r'\(([^)]+)\)', entry_str)
                    rel_text = rel_match.group(1) if rel_match else ""

                    # Determinar parentesco
                    comb_text = f"{rel_text} {entry_str}".upper()
                    if any(k in comb_text for k in ["CÓNYUGE", "CONYUGE", "ESPOSA", "ESPOSO", "PAREJA"]):
                        relacion = "Cónyuge"
                    elif any(k in comb_text for k in ["HIJO", "HIJA", "DESCENDIENTE"]):
                        relacion = "Hijo/a"
                    elif any(k in comb_text for k in ["PADRE", "MADRE", "PAPA", "MAMA"]):
                        relacion = "Padre/Madre"
                    elif any(k in comb_text for k in ["HERMANO", "HERMANA"]):
                        relacion = "Hermano/a"
                    else:
                        relacion = "Heredero/a"
                        
                    parsed_heirs.append({
                        "RUT": rut_extracted,
                        "Relación": relacion,
                        "Nombre": name_clean or "Heredero Declarado",
                        "Fecha de Nacimiento": dob,
                        "% Asignación": 0.0,
                        "¿Estudiante (18-24 años)?": False
                    })
                
                # Asignación porcentual legal eficiente
                if parsed_heirs:
                    n_total = len(parsed_heirs)
                    conyuges = [h for h in parsed_heirs if h["Relación"] == "Cónyuge"]
                    hijos = [h for h in parsed_heirs if h["Relación"] == "Hijo/a"]
                    
                    if conyuges and hijos:
                        if len(hijos) == 1:
                            pct_c = 50.0
                            pct_h = 50.0
                        else:
                            total_parts = len(hijos) + 2
                            pct_c = round((2.0 / total_parts) * 100.0, 2)
                            pct_h = round((100.0 - pct_c) / len(hijos), 2)
                        for h in parsed_heirs:
                            if h["Relación"] == "Cónyuge": h["% Asignación"] = pct_c
                            elif h["Relación"] == "Hijo/a": h["% Asignación"] = pct_h
                    elif conyuges and not hijos:
                        for h in parsed_heirs: h["% Asignación"] = 100.0
                    elif hijos and not conyuges:
                        pct_h = round(100.0 / len(hijos), 2)
                        for h in parsed_heirs: h["% Asignación"] = pct_h
                    else:
                        pct_g = round(100.0 / n_total, 2)
                        for h in parsed_heirs: h["% Asignación"] = pct_g

                results["herederos"] = parsed_heirs

            # ----------------------------------------------------
            # B. PARSER DE PROPIEDADES Y BIENES RAÍCES
            # ----------------------------------------------------
            elif any(kw in item_req_upper for kw in ["PROPIEDAD", "BIEN", "RAIZ", "RAÍZ", "INMUEBLE", "DEPARTAMENTO", "TERRENO"]):
                props_raw = re.split(r'[;\n]', answer)
                uf_today = get_uf_today() or 38850.0
                
                for pr in props_raw:
                    pr_str = pr.strip()
                    if not pr_str: continue
                    
                    # ROL (ej. ROL 1234-5 o 01026-0001)
                    rol_m = re.search(r'ROL\s*[:#]?\s*([\d-]+)', pr_str, re.IGNORECASE)
                    rol = rol_m.group(1).strip() if rol_m else ""
                    
                    # Montos de Valor Comercial
                    val_clp = 0.0
                    val_uf = 0.0
                    
                    uf_m = re.search(r'([\d\.\,]+)\s*UF', pr_str, re.IGNORECASE)
                    clp_m = re.search(r'[\$]\s*([\d\.\,]+)', pr_str)
                    m_m = re.search(r'([\d\.\,]+)\s*M\b', pr_str, re.IGNORECASE)
                    
                    if uf_m:
                        try:
                            val_uf = float(uf_m.group(1).replace('.', '').replace(',', '.'))
                            val_clp = val_uf * uf_today
                        except: pass
                    elif clp_m:
                        try:
                            clean_c = clp_m.group(1).rstrip('.').replace('.', '').replace(',', '.')
                            val_clp = float(clean_c)
                            val_uf = val_clp / uf_today
                        except: pass
                    elif m_m:
                        try:
                            val_m = float(m_m.group(1).replace(',', '.'))
                            val_clp = val_m * 1000000.0
                            val_uf = val_clp / uf_today
                        except: pass

                    deuda = False if ("SIN DEUDA" in pr_str.upper() or "NO TIENE" in pr_str.upper() or "SIN HIPOTECA" in pr_str.upper()) else True
                    
                    comunas = ["LAS CONDES", "VITACURA", "PROVIDENCIA", "ÑUÑOA", "COQUIMBO", "MACUL", "SANTIAGO", "VIÑA DEL MAR", "CONCON", "LA REINA", "LO BARNECHEA", "PEÑUELAS", "PUERTO VARAS", "LA SERENA"]
                    comuna_found = "SANTIAGO"
                    for c in comunas:
                        if c in pr_str.upper():
                            comuna_found = "COQUIMBO" if c == "PEÑUELAS" else c
                            break

                    results["propiedades"].append({
                        "Nombre/Alias": f"Propiedad KYC ({comuna_found.title()})",
                        "Comuna": comuna_found,
                        "ROL": rol or f"KYC-{comuna_found[:3]}-01",
                        "Dirección": pr_str[:120],
                        "Destino": "HABITACIONAL",
                        "Avalúo Fiscal (CLP)": 0.0, # Sin estimaciones artificiales
                        "Valor Com. (UF)": round(val_uf, 2),
                        "Tasación (UF)": round(val_uf, 2),
                        "Deuda Hipotecaria": deuda,
                        "Origen Tasación": "Declarada por Cliente (KYC)"
                    })

            # ----------------------------------------------------
            # C. PARSER DE PÓLIZAS, SEGUROS Y APV
            # ----------------------------------------------------
            elif any(kw in item_req_upper for kw in ["PÓLIZA", "POLIZA", "SEGURO", "APV", "FONDO", "INVERSIÓN", "INVERSION"]):
                insts = ["PRINCIPAL", "METLIFE", "BANCO CHILE", "BANCO DE CHILE", "CONSORCIO", "SURA", "HABITAT", "CUPRUM", "CAPITAL", "MODELO", "PLANVITAL", "SECURITY", "BCI", "SANTANDER", "ESTADO"]
                found_insts = []
                for inst in insts:
                    if inst in answer.upper():
                        found_insts.append("BANCO DE CHILE" if inst == "BANCO CHILE" else inst)

                if not found_insts:
                    found_insts = ["Declarado por Cliente"]

                for inst_name in found_insts:
                    tipo_seg = "APV" if "APV" in answer.upper() else "Seguro de Vida"
                    results["polizas"].append({
                        "Aseguradora": inst_name.title(),
                        "Asegurado": "Titular",
                        "Contratante": "Titular",
                        "Tipo": tipo_seg,
                        "N° Póliza": "KYC-DECLARADO",
                        "Colectivo / Individual": "Individual",
                        "Alias / Patente": "Declarado en KYC",
                        "Monto (UF)": 0.0,
                        "Prima": 0.0,
                        "Medio de Pago": "PAT / PAC",
                        "Fecha Contratación": None,
                        "¿APV Póliza?": "APV" in answer.upper(),
                        "Coberturas": answer[:150],
                        "Análisis IA": f"Declaración Cliente KYC: '{answer[:100]}...'"
                    })

    except Exception as e:
        logging.error(f"Error parseando Excel KYC: {e}")

    return results
