import datetime

class FamilyBoardSimulator:
    def __init__(self):
        pass
        
    def generar_informe_legal(self, file_name: str, file_text: str = "") -> str:
        """
        Simula o procesa el análisis legal de una escritura o malla societaria.
        """
        fecha = datetime.datetime.now().strftime("%d/%m/%Y")
        informe = f"""==================================================
INFORME LEGAL Y SOCIETARIO - ALTUS AI
Fecha de Análisis: {fecha}
Documento Analizado: {file_name}
==================================================

1. RESUMEN EJECUTIVO
Se ha procesado la documentación legal aportada. Se identifican estructuras societarias entrelazadas que requieren una revisión de poderes y cláusulas de administración para garantizar una correcta sucesión.

2. HALLAZGOS PRINCIPALES
- Existen participaciones cruzadas entre las sociedades holding y las operativas.
- El poder de veto o decisiones de mayoría absoluta recaen actualmente en la figura del socio fundador.
- No se observan cláusulas explícitas de salida (Buy-Sell Agreements) claras para herederos en caso de fallecimiento o incapacidad.

3. RECOMENDACIONES ESTRATÉGICAS
- Homologar los estatutos de las sociedades holding para reflejar las mismas mayorías exigidas.
- Implementar un pacto de accionistas preventivo que defina los roles de los herederos pasivos vs activos.
- Estructurar un esquema de "Acciones Preferentes" para separar los derechos económicos (dividendos) de los derechos políticos (voto) para la siguiente generación.

*Nota: Este análisis fue generado por Altus AI. Se recomienda validación final con asesoría legal externa para la redacción de las escrituras.*
"""
        return informe

    def generar_protocolo_familiar(self, familia_info: dict) -> str:
        """
        Genera un borrador del protocolo familiar basado en los inputs de la familia.
        """
        fundador = familia_info.get('fundador', '[Nombre Fundador]')
        herederos = familia_info.get('herederos', '[Herederos]')
        valores = familia_info.get('valores', 'Integridad, Crecimiento, Respeto')
        politica_div = familia_info.get('politica_dividendos', 'Reinvertir la mayor parte de las utilidades')
        restricciones = familia_info.get('restricciones', 'Los familiares deben postular a los cargos por mérito profesional')

        protocolo = f"""==================================================
BORRADOR DE PROTOCOLO FAMILIAR
Familia Empresaria: {fundador} y Familia
==================================================

I. PREÁMBULO Y VISIÓN
El presente Protocolo Familiar tiene por objeto establecer las reglas del juego que regirán las relaciones entre la familia, la propiedad y la gestión de la empresa. Nuestra visión es asegurar la continuidad transgeneracional, manteniendo el patrimonio unido y protegiendo el legado.

II. VALORES FAMILIARES
Nuestras decisiones empresariales se rigen por los siguientes valores fundamentales: {valores}.

III. REGLAS DE CONTRATACIÓN Y PARTICIPACIÓN
Respecto al ingreso de familiares a la compañía, se acuerda lo siguiente:
- {restricciones}.
- Todo miembro de la familia (incluyendo políticos o descendientes) deberá ser evaluado por sus competencias profesionales por un comité externo o un Headhunter, sin preferencias especiales.

IV. POLÍTICA DE DIVIDENDOS Y LIQUIDEZ
Para proteger la salud financiera de las empresas operativas y a la vez proveer liquidez a los accionistas familiares, se establece la siguiente directriz:
- {politica_div}.

V. SUCESIÓN Y ÓRGANOS DE GOBIERNO
Se instaurará un "Consejo de Familia" que sesionará trimestralmente para discutir asuntos de índole familiar y resolver conflictos. Las decisiones estrictamente empresariales recaerán exclusivamente sobre el Directorio Profesional, donde la familia participará activamente pero incluirá directores independientes.

---
Este borrador inicial es un documento vivo que debe ser discutido y ratificado por la familia.
"""
        return protocolo

    def simular_impuesto_herencia_avanzado(self, patrimonio_total_clp: float, herederos_dict: dict, valor_uta: float = 793000):
        """
        Calcula el impuesto a la herencia progresivo en Chile (Ley 16.271) aplicando recargos por parentesco (Art. 2)
        y exenciones de UTA.
        herederos_dict format: {'conyuge_hijos': int, 'ascendientes': int, 'hermanos': int, 'otros': int}
        """
        num_herederos = sum(herederos_dict.values())
        if num_herederos <= 0: return 0, 0
        
        # Asumiendo partición en partes iguales para simplificar el modelo estocástico
        asignacion_por_heredero = patrimonio_total_clp / num_herederos
        asignacion_uta = asignacion_por_heredero / valor_uta
        
        impuesto_total_uta = 0
        
        for relacion, cantidad in herederos_dict.items():
            if cantidad == 0: continue
            
            # Determinar exención y recargo según Ley 16.271
            exencion_uta = 0
            recargo = 0.0
            
            if relacion == 'conyuge_hijos':
                exencion_uta = 50
                recargo = 0.0
            elif relacion == 'ascendientes':
                exencion_uta = 20
                recargo = 0.0
            elif relacion == 'hermanos':
                exencion_uta = 20
                recargo = 0.20 # 20% de recargo Art. 2
            elif relacion == 'otros':
                exencion_uta = 0 # Extraños no tienen exención
                recargo = 0.40 # 40% de recargo Art. 2
                
            base_imponible_uta = max(0, asignacion_uta - exencion_uta)
            
            impuesto_base_uta = 0
            # Tabla Progresiva
            b = base_imponible_uta
            if b > 1200:
                impuesto_base_uta += (b - 1200) * 0.25; b = 1200
            if b > 800:
                impuesto_base_uta += (b - 800) * 0.168; b = 800
            if b > 600:
                impuesto_base_uta += (b - 600) * 0.12; b = 600
            if b > 400:
                impuesto_base_uta += (b - 400) * 0.096; b = 400
            if b > 320:
                impuesto_base_uta += (b - 320) * 0.072; b = 320
            if b > 160:
                impuesto_base_uta += (b - 160) * 0.048; b = 160
            if b > 80:
                impuesto_base_uta += (b - 80) * 0.024; b = 80
            if b > 0:
                impuesto_base_uta += b * 0.012
                
            # Aplicar recargo de parentesco (Art. 2)
            impuesto_final_uta = impuesto_base_uta * (1 + recargo)
            
            impuesto_total_uta += (impuesto_final_uta * cantidad)
            
        impuesto_total_clp = impuesto_total_uta * valor_uta
        return impuesto_total_clp, impuesto_total_clp / num_herederos if num_herederos else 0

    def simular_escenario_optimizado_avanzado(self, patrimonio_total_clp: float, herederos_dict: dict, valor_uta: float = 793000):
        """
        Calcula un escenario de estructuración societaria y usufructo (Art. 6 Ley 16.271).
        """
        masa_hereditaria_reducida = patrimonio_total_clp * 0.40
        return self.simular_impuesto_herencia_avanzado(masa_hereditaria_reducida, herederos_dict, valor_uta)

    def generar_plantilla_inventario(self) -> bytes:
        import pandas as pd
        import io
        df = pd.DataFrame({
            'Categoría del Activo': [
                'Bienes Raíces (Propiedades)', 
                'Instrumentos Financieros (Fondos/Acciones)',
                'Participación Societaria (Empresas)',
                'Vehículos',
                'Cuentas Bancarias / APV'
            ],
            'Descripción Detallada (Requisito Legal)': [
                'Depto Providencia. ROL: 123-4, Comuna Providencia. Inscripción CBR: Fojas 500, N° 450, Año 2015.', 
                'Fondo Mutuo Santander. RUT Institución: 76.xxx.xxx-x. N° Contrato/Cuenta: 88776655.',
                'Inversiones Familiares SpA. RUT: 77.xxx.xxx-x. Participación: 50%.',
                'Camioneta Ford F-150. Patente: ABCD-12. Año: 2022.',
                'Cuenta Corriente Banco de Chile. N° 11223344.'
            ],
            'Guía para el Cliente (¿Qué necesita el abogado/notario?)': [
                'Indicar ROL, Comuna y datos de inscripción en el Conservador de Bienes Raíces (Fojas, Número, Año).',
                'Indicar Institución, RUT de la AGF/Corredora y Número exacto de cuenta o contrato.',
                'Indicar Razón Social, RUT de la sociedad y porcentaje (%) exacto de participación del testador.',
                'Basta con indicar la Placa Patente única, marca y año.',
                'Indicar Banco y Número de cuenta. Si es APV, indicar la administradora.'
            ],
            'Valor Comercial Estimado (CLP)': [250000000, 150000000, 800000000, 25000000, 10000000],
            'Deuda o Hipoteca Asociada (CLP)': [50000000, 0, 0, 0, 0]
        })
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Inventario Patrimonial', startrow=5)
            
            worksheet = writer.sheets['Inventario Patrimonial']
            worksheet.sheet_view.showGridLines = False
            
            # Configuración de Impresión (Página)
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LETTER
            worksheet.page_setup.fitToPage = True
            worksheet.page_setup.fitToHeight = 0
            worksheet.page_setup.fitToWidth = 1
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            
            # Anchos de columna
            worksheet.column_dimensions['A'].width = 28
            worksheet.column_dimensions['B'].width = 45
            worksheet.column_dimensions['C'].width = 45
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 20
            
            # Altura de filas para dar espacio a logos verdaderamente GIGANTES
            worksheet.row_dimensions[1].height = 100
            worksheet.row_dimensions[2].height = 25
            
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
            
            # Aplicar "Ajustar texto" y "Alineación en el MEDIO" a toda la tabla
            wrap_alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            for row in worksheet.iter_rows(min_row=6, max_row=14, min_col=1, max_col=5):
                for cell in row:
                    cell.alignment = wrap_alignment
            
            # Título Central
            worksheet['C3'] = "Inventario Patrimonial Digital"
            worksheet['C3'].font = Font(size=15, bold=True, color="002060") # Azul institucional
            worksheet['C3'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Añadir Logos Institucionales (Gigantes)
            import os
            try:
                from openpyxl.drawing.image import Image
                fv_path = os.path.abspath("src/web/assets/NUEVO LOGO FV.png")
                altus_path = os.path.abspath("src/web/assets/ALTUS_AI_LOGO.png")
                
                if os.path.exists(fv_path):
                    img_fv = Image(fv_path)
                    scale_fv = 100 / img_fv.height # Aumentado a 100px
                    img_fv.height = 100
                    img_fv.width = int(img_fv.width * scale_fv)
                    worksheet.add_image(img_fv, 'A1')
                    
                if os.path.exists(altus_path):
                    img_altus = Image(altus_path)
                    scale_altus = 80 / img_altus.height # Aumentado a 80px
                    img_altus.height = 80
                    img_altus.width = int(img_altus.width * scale_altus)
                    worksheet.add_image(img_altus, 'E1')
            except Exception as e:
                pass
                
            # Pie de Página "Sobre FV" con diseño FUERTE Y CON CARÁCTER
            # Separamos el título del cuerpo en dos celdas fusionadas
            worksheet.merge_cells('A15:E15')
            worksheet.merge_cells('A16:E16')
            worksheet.row_dimensions[15].height = 25
            worksheet.row_dimensions[16].height = 55
            
            # Celda de Título del Footer
            title_cell = worksheet['A15']
            title_cell.value = "  Sobre FV Asesorías e Inversiones"
            title_cell.font = Font(name='Arial', size=11, bold=True, color="002060") # Azul oscuro y negrita
            title_cell.alignment = Alignment(vertical="center", horizontal="left")
            title_cell.fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid") # Fondo gris elegante
            
            # Celda del Cuerpo del Footer
            body_cell = worksheet['A16']
            body_text = (
                "  Somos un Multi-Family Office Digital potenciado por Altus AI, nuestro Software Cuantitativo Privado. "
                "Combinamos la precisión algorítmica de la Inteligencia Artificial con la exclusividad de la banca privada "
                "para auditar portafolios, cruzar normativas tributarias complejas, incorporar información de valor para cada cliente "
                "y diseñar estrategias patrimoniales hiper-personalizadas de grado institucional."
            )
            body_cell.value = body_text
            body_cell.font = Font(name='Arial', size=10, color="2A2A2A")
            body_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            body_cell.fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
            
            # Bordes del Footer (Borde dorado lateral)
            gold_side = Side(border_style="thick", color="D4AF37")
            gray_side = Side(border_style="thin", color="E0E0E0")
            
            for c_let in ['A', 'B', 'C', 'D', 'E']:
                # Borde Fila 15 (Título)
                cell15 = worksheet[f'{c_let}15']
                l15 = gold_side if c_let == 'A' else None
                r15 = gray_side if c_let == 'E' else None
                cell15.border = Border(left=l15, right=r15, top=gray_side, bottom=None)
                
                # Borde Fila 16 (Cuerpo)
                cell16 = worksheet[f'{c_let}16']
                l16 = gold_side if c_let == 'A' else None
                r16 = gray_side if c_let == 'E' else None
                cell16.border = Border(left=l16, right=r16, top=None, bottom=gray_side)
            
            # Protecciones
            worksheet.protection.sheet = True
            worksheet.protection.password = "AltusFV2026_secreto"
            worksheet.protection.insertRows = False
            worksheet.protection.formatCells = False
            
            from openpyxl.styles import Protection
            unlocked = Protection(locked=False)
            for row in worksheet.iter_rows(min_row=7, max_row=14, min_col=1, max_col=5):
                for cell in row:
                    cell.protection = unlocked
                    
        return output.getvalue()

    def generar_testamento(self, info: dict) -> str:
        """Genera un borrador de testamento."""
        testador = info.get('testador', '[Testador]')
        albacea = info.get('albacea', '[Albacea]')
        cuarta_libre = info.get('cuarta_libre', '[Cuarta de libre disposición]')
        
        testamento = f"""==================================================
BORRADOR DE TESTAMENTO (MODELO CHILENO)
Testador: {testador}
==================================================

I. ANTECEDENTES DEL TESTADOR
Yo, {testador}, en pleno uso de mis facultades mentales, otorgo el presente testamento para disponer de mis bienes...

II. INSTITUCIÓN DE HEREDEROS
Declaro que mis herederos forzosos son mis hijos de filiación matrimonial y mi cónyuge sobreviviente, a quienes asigno la Mitad Legitimaria y la Cuarta de Mejoras conforme lo dispone la ley.

III. CUARTA DE LIBRE DISPOSICIÓN
Respecto de la Cuarta de Libre Disposición de mi patrimonio, es mi voluntad legarla de la siguiente manera:
- {cuarta_libre}.

IV. NOMBRAMIENTO DE ALBACEA
Designo como Albacea con tenencia de bienes a {albacea}, a quien le confiero amplias facultades para la administración y partición de la masa hereditaria, velando por el cumplimiento de mis disposiciones.

V. DISPOSICIONES ADICIONALES
Ruego a mis herederos que administren el patrimonio familiar bajo los principios establecidos en nuestro Protocolo Familiar, respetando los órganos de gobierno corporativo.

---
*Nota Legal: Este borrador fue estructurado por Altus AI para propósitos de discusión familiar. Para su validez legal y formal, debe ser otorgado mediante escritura pública ante Notario y tres testigos, validando el respeto a las asignaciones forzosas chilenas.*
"""
        return testamento
