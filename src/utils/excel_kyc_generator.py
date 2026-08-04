import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generar_excel_kyc_corporativo(client_name="Cliente", missing_herederos=True, missing_propiedades=True, missing_polizas=True, custom_items=None):
    """
    Genera un archivo Excel con formato corporativo de alta calidad (Altus AI / FV Asesorías)
    para el Formulario de Onboarding & Auditoría Patrimonial (KYC).
    
    Incluye 3 Hojas:
    1. Formulario KYC 360° (Con columna de respuesta destacada en amarillo/azul)
    2. Guía - Dónde Obtener tus Datos (Instrucciones paso a paso para el cliente)
    3. Seguridad y Marco Legal (Garantías de cifrado y resguardo Ley N° 19.628)
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # ESTILOS CORPORATIVOS ELEGANTES
    # ----------------------------------------------------
    title_fill = PatternFill(start_color="0A2342", end_color="0A2342", fill_type="solid") # Dark Navy
    title_font = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
    
    sub_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate Dark
    sub_font = Font(name="Calibri", color="CBD5E1", italic=True, size=10.5)
    
    warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Amber / Gold Warning
    warn_font = Font(name="Calibri", color="78350F", bold=True, size=10.5)
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Charcoal / Dark Slate
    header_font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    
    action_header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid") # Sky Blue Action Header
    action_header_font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    
    label_font = Font(name="Calibri", bold=True, size=11, color="0F172A")
    why_font = Font(name="Calibri", size=10, color="334155")
    example_font = Font(name="Calibri", italic=True, size=9.5, color="0369A1") # Blueish tint for examples
    answer_font = Font(name="Calibri", size=11, color="0F172A")
    
    row_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_fill_odd = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    answer_cell_fill = PatternFill(start_color="FFFDF0", end_color="FFFDF0", fill_type="solid") # Soft yellow highlight for input area
    
    thin_border = Border(
        left=Side(style='thin', color="CBD5E1"),
        right=Side(style='thin', color="CBD5E1"),
        top=Side(style='thin', color="CBD5E1"),
        bottom=Side(style='thin', color="CBD5E1")
    )

    # ----------------------------------------------------
    # HOJA 1: Formulario KYC 360°
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Formulario KYC 360"
    ws1.views.sheetView[0].showGridLines = False
    
    ws1.merge_cells('A1:D1')
    clean_name = client_name.upper() if client_name else "CLIENTE"
    ws1['A1'] = f"FORMULARIO DE ONBOARDING & AUDITORIA PATRIMONIAL (KYC) - {clean_name}"
    ws1['A1'].fill = title_fill
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 38
    
    ws1.merge_cells('A2:D2')
    ws1['A2'] = "Altus AI | Family Office & Gestion Patrimonial Integrada - Proceso de Recopilacion de Antecedentes"
    ws1['A2'].fill = sub_fill
    ws1['A2'].font = sub_font
    ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 24
    
    ws1.merge_cells('A3:D3')
    ws1['A3'] = "⚠️ INSTRUCCION IMPORTANTE: En la columna 'Respuesta / Detalle del Cliente' debe INGRESAR LOS DATOS ESPECIFICOS solicitados segun la Pestaña 2 (Guía). No responda unicamente 'Sí' o 'De acuerdo'."
    ws1['A3'].fill = warn_fill
    ws1['A3'].font = warn_font
    ws1['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.row_dimensions[3].height = 36
    
    headers = [
        "Dato Requerido por Altus AI",
        "¿Por qué lo necesitamos?",
        "Ejemplo / Formato Sugerido",
        "Respuesta / Detalle del Cliente (Escribir aquí)"
    ]
    cols = ['A', 'B', 'C', 'D']
    
    for col, h in zip(cols, headers):
        cell = ws1[f'{col}4']
        cell.value = h
        if col == 'D':
            cell.fill = action_header_fill
            cell.font = action_header_font
        else:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.row_dimensions[4].height = 34
    
    kyc_items = []
    if custom_items:
        kyc_items = custom_items
    else:
        if missing_herederos:
            kyc_items.append((
                "Detalle de Herederos (Parentesco, RUT, Fecha de Nacimiento y Composición Familiar)",
                "Requerido para la planificación de sucesión, cálculo de exenciones tributarias de herencia (Ley N° 16.271) y tabla actuarial.",
                "Escriba Nombres, Parentesco, RUT y Fecha de Nacimiento. Ej: María González (Cónyuge; RUT 9.017.656-0; Fecha Nac: 15/05/1955), Pedro Pérez (Hijo; RUT 10.870.820-4; Fecha Nac: 20/10/1985)."
            ))
        if missing_propiedades:
            kyc_items.append((
                "Detalle de Propiedades y Bienes Raíces",
                "Necesario para calcular patrimonio neto consolidado, avalúo comercial en UF y planificar liquidez sucesoria.",
                "Indique ROL y comuna. Ej: ROL 1234-5 Las Condes, Valor comercial ~10.000 UF, Deuda Hipotecaria $50M Banco de Chile."
            ))
        if missing_polizas:
            kyc_items.append((
                "Pólizas de Seguro de Vida o APV",
                "Para auditar estructuras inembargables y maximizar beneficios tributarios (Art. 57 LIR, Art. 42 bis LIR).",
                "Detalle aseguradora e institución. Ej: Principal APV Régimen B $20M, MetLife Seguro de Vida 5.000 UF."
            ))
        
        if not kyc_items:
            kyc_items = [
                ("Nombres, Apellidos y RUT del Titular", "Identificación legal del titular para la elaboración del informe patrimonial.", "Ej: Nelson Moraga Benavides | RUT: 12.345.678-9 | Fecha Nac: 15/08/1980."),
                ("Detalle de Herederos, RUT y Fecha de Nacimiento", "Planificación de sucesión y cálculo de exenciones tributarias de herencia.", "Ej: María González (Cónyuge; RUT 9.017.656-0; Fecha Nac: 15/05/1955), Pedro Pérez (Hijo; RUT 10.870.820-4; Fecha Nac: 20/10/1985)."),
                ("Bienes Raíces (ROL, Comuna y Deudas)", "Consolidación patrimonial en UF y auditoría de garantías.", "Ej: ROL 1234-5 Las Condes, ~10.000 UF, Hipoteca $50M."),
                ("Pólizas de Vida / APV (Régimen A/B)", "Análisis de inembargabilidad y franquicias tributarias (Art. 57 LIR).", "Ej: Principal APV Régimen B ($25M), MetLife Seguro Vida (5.000 UF).")
            ]
            
    row = 5
    for idx, (item, why, example) in enumerate(kyc_items):
        r_fill = row_fill_odd if idx % 2 == 1 else row_fill_even
        
        ws1[f'A{row}'] = item
        ws1[f'A{row}'].font = label_font
        ws1[f'A{row}'].fill = r_fill
        ws1[f'A{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws1[f'A{row}'].border = thin_border
        
        ws1[f'B{row}'] = why
        ws1[f'B{row}'].font = why_font
        ws1[f'B{row}'].fill = r_fill
        ws1[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws1[f'B{row}'].border = thin_border
        
        ws1[f'C{row}'] = example
        ws1[f'C{row}'].font = example_font
        ws1[f'C{row}'].fill = r_fill
        ws1[f'C{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws1[f'C{row}'].border = thin_border
        
        ws1[f'D{row}'] = ""
        ws1[f'D{row}'].font = answer_font
        ws1[f'D{row}'].fill = answer_cell_fill
        ws1[f'D{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws1[f'D{row}'].border = thin_border
        
        max_len = max(len(item), len(why), len(example))
        ws1.row_dimensions[row].height = max(55, min(95, int(max_len / 2.2)))
        row += 1

    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 46
    ws1.column_dimensions['C'].width = 46
    ws1.column_dimensions['D'].width = 52

    # ----------------------------------------------------
    # HOJA 2: Guía - Dónde Obtener tus Datos (NUEVA PESTAÑA PRÁCTICA)
    # ----------------------------------------------------
    ws_guide = wb.create_sheet(title="Guía - Dónde Obtener Datos")
    ws_guide.views.sheetView[0].showGridLines = False
    
    ws_guide.merge_cells('A1:C1')
    ws_guide['A1'] = "GUIA PRACTICA: ¿DONDE OBTENER CADA DATO SOLICITADO?"
    ws_guide['A1'].fill = title_fill
    ws_guide['A1'].font = title_font
    ws_guide['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_guide.row_dimensions[1].height = 38
    
    ws_guide.merge_cells('A2:C2')
    ws_guide['A2'] = "Instrucciones de descarga rápida en portales oficiales y documentos habituales"
    ws_guide['A2'].fill = sub_fill
    ws_guide['A2'].font = sub_font
    ws_guide['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_guide.row_dimensions[2].height = 24
    
    headers_guide = ["Dato / Antecedente Solicitado", "¿Dónde encontrarlo? (Fuente u Origen)", "Pasos / Documento Sugerido para Obtención"]
    guide_cols = ['A', 'B', 'C']
    
    for col, h in zip(guide_cols, headers_guide):
        cell = ws_guide[f'{col}3']
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_guide.row_dimensions[3].height = 30
    
    guide_data = [
        (
            "1. Retenciones de Impuesto (2ª Categoría)",
            "Liquidación de Sueldo o Portal SII.cl",
            "• En tu Liquidación de Sueldo: Buscar el ítem 'Impuesto Único de Segunda Categoría' en Descuentos Legales.\n• En sii.cl: Menú 'Mis Declaraciones' > Certificados de Sueldo (F1887)."
        ),
        (
            "2. Intereses Dividendo Hipotecario (Art. 55 bis)",
            "Banco o Mutuaria Hipotecaria",
            "• En la web de tu Banco: Sección Créditos Hipotecarios > Descargar 'Certificado Tributario Art. 55 bis'.\n• Muestra el monto de intereses pagados en el año."
        ),
        (
            "3. Gastos en Educación de Hijos (Art. 55 ter)",
            "Colegio o Universidad del Hijo",
            "• Certificado de Matrícula o Escolaridad emitido por el establecimiento educacional + comprobante de pago de colegiatura."
        ),
        (
            "4. Datos de Cónyuge e Hijos (RUT / Fechas Nac.)",
            "Cédula de Identidad o Registro Civil",
            "• Cédula de Identidad del familiar.\n• Descargar gratis Certificado de Nacimiento / Matrimonio en registrocivil.cl."
        ),
        (
            "5. ROL y Avalúo Comercial de Propiedades",
            "Portal SII.cl o Recibo de Contribuciones",
            "• En sii.cl: Menú 'Bienes Raíces' > 'Mis Propiedades' (aparece el número de ROL y comuna).\n• O revisar el aviso de pago de Contribuciones."
        ),
        (
            "6. APV Actual y Pólizas de Seguro",
            "Aseguradora, AGF o AFP Actual",
            "• Cartola anual o resumen de saldo descargable desde el sitio cliente de la institución (ej: Principal, MetLife, BICE, LarrainVial, etc.)."
        )
    ]
    
    row = 4
    for idx, (item, source, steps) in enumerate(guide_data):
        r_fill = row_fill_odd if idx % 2 == 1 else row_fill_even
        
        ws_guide[f'A{row}'] = item
        ws_guide[f'A{row}'].font = label_font
        ws_guide[f'A{row}'].fill = r_fill
        ws_guide[f'A{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws_guide[f'A{row}'].border = thin_border
        
        ws_guide[f'B{row}'] = source
        ws_guide[f'B{row}'].font = Font(name="Calibri", bold=True, size=10, color="0369A1")
        ws_guide[f'B{row}'].fill = r_fill
        ws_guide[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws_guide[f'B{row}'].border = thin_border
        
        ws_guide[f'C{row}'] = steps
        ws_guide[f'C{row}'].font = why_font
        ws_guide[f'C{row}'].fill = r_fill
        ws_guide[f'C{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws_guide[f'C{row}'].border = thin_border
        
        ws_guide.row_dimensions[row].height = 55
        row += 1
        
    ws_guide.column_dimensions['A'].width = 38
    ws_guide.column_dimensions['B'].width = 35
    ws_guide.column_dimensions['C'].width = 75

    # ----------------------------------------------------
    # HOJA 3: Seguridad y Marco Legal
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Seguridad y Marco Legal")
    ws2.views.sheetView[0].showGridLines = False
    
    ws2.merge_cells('A1:B1')
    ws2['A1'] = "PROTECCION DE DATOS Y VALIDEZ LEGAL"
    ws2['A1'].fill = title_fill
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 38
    
    ws2.merge_cells('A2:B2')
    ws2['A2'] = "Garantias de confidencialidad de FV Asesorias e Inversiones & Altus AI"
    ws2['A2'].fill = sub_fill
    ws2['A2'].font = sub_font
    ws2['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[2].height = 24
    
    info_legal = [
        ("🔒 Encriptacion y Seguridad Bancaria", "Sus datos estan protegidos bajo cifrado AES-256 en reposo y TLS 1.3 en transito, garantizando confidencialidad bajo Secreto Patrimonial."),
        ("🛡️ Ley N° 19.628 (Vida Privada)", "Sus antecedentes son de uso exclusivo del equipo asignado a su asesoria MFO. No se comparten ni comercializan con terceros bajo ninguna circunstancia."),
        ("⚖️ Firma Electronica (Ley N° 19.799)", "Conforme a la Ley N° 19.799 sobre Documentos Electronicos en Chile, el envio de este formulario desde su casilla de correo personal constituye una Firma Electronica Simple (FES) valida.")
    ]
    
    row = 4
    for title, desc in info_legal:
        ws2[f'A{row}'] = title
        ws2[f'A{row}'].font = Font(name="Calibri", bold=True, size=11, color="0A2342")
        ws2[f'A{row}'].border = thin_border
        ws2[f'A{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws2[f'A{row}'].fill = row_fill_odd
        
        ws2[f'B{row}'] = desc
        ws2[f'B{row}'].font = Font(name="Calibri", size=10.5, color="334155")
        ws2[f'B{row}'].border = thin_border
        ws2[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws2[f'B{row}'].fill = row_fill_even
        
        ws2.row_dimensions[row].height = 45
        row += 1
        
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 90

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generar_excel_apv_reliquidacion(client_name="José González Daza"):
    """
    Genera un Excel específicamente estructurado para capturar los datos requeridos por los simuladores
    'APV Inteligente' y 'Reliquidación de Impuestos IGC' (Cónyuge, Hijos, Fechas Nac, Renta Bruta/Líquida, etc.).
    """
    custom_items = [
        (
            "1. Datos de la Cónyuge (Nombre, RUT, Fecha Nac., Ingresos)",
            "Requerido para evaluar el régimen de tributación conjunta o independiente, exenciones del Impuesto a las Herencias y planificación de liquidez.",
            "Ej: Ver Cédula de Identidad o registrocivil.cl. Ej: María Elena Pérez | RUT: 13.456.789-0 | Fecha Nac: 12/04/1982 | Sueldo líquido ~$1.8M."
        ),
        (
            "2. Detalle de Hijos (Nombres, RUT, Fechas Nac., Nivel Educacional)",
            "Necesario para calcular el Crédito Tributario por Gastos en Educación (Art. 55 ter LIR) y planificar distribución sucesoria de legitimarios.",
            "Ej: Ver Cédula o Colegio. Ej: Mateo González (15/08/2012, Colegio), Sofía González (03/11/2016, Colegio)."
        ),
        (
            "3. Renta Bruta / Sueldo Líquido Mensual Aproximado",
            "Fundamental para alimentar el Simulador APV Inteligente y definir el tramo exacto del Impuesto Global Complementario (IGC) entre 4% y 40%.",
            "Ej: Ver en Liquidación de Sueldo mensual. Ej: Sueldo líquido mensual ~$3.500.000 (Renta imponible bruta ~$4.200.000)."
        ),
        (
            "4. Retenciones de Impuesto 2ª Categoría o Créditos Vigentes",
            "Requerido por el Simulador de Reliquidación IGC para proyectar la devolución anual de impuestos a favor del cliente.",
            "Ej: Ver en Liquidación de Sueldo ('Impuesto Único') o Banco ('Certificado 55 bis'). Ej: Retención ~$350k/mes | Dividendo $400k/mes."
        ),
        (
            "5. Monto Objetivo de Aporte Mensual APV & Régimen Deseado",
            "Permite comparar en el simulador el Régimen A (Bonificación estatal del 15%) vs Régimen B (Rebaja directa en la base imponible del IGC).",
            "Ej: Quiero aportar $200.000 mensuales | Deseo evaluar qué régimen me da mayor rentabilidad fiscal."
        ),
        (
            "6. APV Actual Vigente (Institución, Régimen y Saldo Aprox.)",
            "Para auditar comisiones de administración, optimizar rescates defensivos y consolidar la estrategia patrimonial.",
            "Ej: Ver en cartola cliente de tu institución. Ej: Tengo APV en Principal | Régimen B | Saldo acumulado ~$8.500.000."
        )
    ]
    return generar_excel_kyc_corporativo(client_name=client_name, custom_items=custom_items)
